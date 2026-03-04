import os, io, re, json, tempfile, urllib.parse
from flask import Flask, request, jsonify, render_template, send_file, session
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage
import requests
from bs4 import BeautifulSoup
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'dlx-offer-tool-dev-key-change-in-prod')
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max

UPLOAD_FOLDER = tempfile.mkdtemp()
ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv', 'tsv', 'pdf'}

TEMPLATE_COLUMNS = ['Image', 'Item Name', 'Expiration', 'UPC/Item #', 'Quantity', 'Casepack', 'Cost', 'Retail Link', 'FOB']


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# ── File Parsing ──────────────────────────────────────────────────────────────
def parse_upload(filepath, filename):
    ext = filename.rsplit('.', 1)[1].lower()
    if ext == 'csv':
        df = pd.read_csv(filepath, dtype=str)
    elif ext == 'tsv':
        df = pd.read_csv(filepath, sep='\t', dtype=str)
    elif ext in ('xlsx', 'xls'):
        df = parse_excel(filepath)
    elif ext == 'pdf':
        df = parse_pdf(filepath)
    else:
        raise ValueError(f'Unsupported file type: {ext}')
    df = df.fillna('')
    df.columns = [str(c).strip() for c in df.columns]
    return df


def parse_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    data = []
    headers = []
    hyperlinks = {}

    for cell in ws[1]:
        headers.append(str(cell.value or f'Column_{cell.column}').strip())

    # Extract embedded images and save them to temp files, mapped by (row, col)
    image_map = {}
    if hasattr(ws, '_images'):
        for img_idx, img in enumerate(ws._images):
            try:
                anchor = img.anchor
                # Get row/col from anchor — handle different anchor types
                row = None
                col = None
                if hasattr(anchor, '_from'):
                    row = anchor._from.row + 1  # 0-indexed to 1-indexed
                    col = anchor._from.col
                elif hasattr(anchor, 'pos'):
                    # AbsoluteAnchor — estimate row from position
                    pass
                if row is None:
                    print(f'Image {img_idx}: could not determine row/col from anchor type {type(anchor).__name__}')
                    continue

                # Extract image data — try multiple approaches
                img_data = None

                # Approach 1: _data() method (most common in openpyxl)
                if hasattr(img, '_data') and callable(img._data):
                    try:
                        img_data = img._data()
                    except Exception:
                        pass

                # Approach 2: ref attribute (file-like object)
                if not img_data and hasattr(img, 'ref'):
                    try:
                        if hasattr(img.ref, 'read'):
                            img_data = img.ref.read()
                            img.ref.seek(0)
                        elif isinstance(img.ref, bytes):
                            img_data = img.ref
                    except Exception:
                        pass

                # Approach 3: _blob attribute
                if not img_data and hasattr(img, '_blob'):
                    try:
                        img_data = img._blob
                    except Exception:
                        pass

                # Approach 4: Use PIL to re-export if we have a path
                if not img_data and hasattr(img, 'path'):
                    try:
                        from PIL import Image as PILImage
                        from zipfile import ZipFile
                        # Images are stored inside the xlsx (which is a zip)
                        with ZipFile(filepath, 'r') as zf:
                            img_path_in_zip = img.path.lstrip('/')
                            if img_path_in_zip.startswith('xl/'):
                                pass
                            else:
                                img_path_in_zip = 'xl/' + img_path_in_zip
                            try:
                                img_data = zf.read(img_path_in_zip)
                            except KeyError:
                                # Try without xl/ prefix
                                try:
                                    img_data = zf.read(img.path.lstrip('/'))
                                except KeyError:
                                    pass
                    except Exception:
                        pass

                if img_data and len(img_data) > 100:
                    # Determine format from first bytes
                    fmt = 'png'
                    if img_data[:3] == b'\xff\xd8\xff':
                        fmt = 'jpeg'
                    elif img_data[:4] == b'\x89PNG':
                        fmt = 'png'
                    elif img_data[:4] == b'GIF8':
                        fmt = 'gif'
                    elif img_data[:4] == b'RIFF':
                        fmt = 'webp'
                    elif img_data[:4] == b'<svg' or b'<svg' in img_data[:100]:
                        fmt = 'svg'

                    # Convert webp/svg to png for compatibility
                    if fmt in ('webp', 'svg'):
                        try:
                            from PIL import Image as PILImage
                            pil_img = PILImage.open(io.BytesIO(img_data))
                            buf = io.BytesIO()
                            pil_img.save(buf, format='PNG')
                            img_data = buf.getvalue()
                            fmt = 'png'
                        except Exception:
                            pass

                    img_filename = f'extracted_{img_idx}_{row}_{col}.{fmt}'
                    img_path = os.path.join(UPLOAD_FOLDER, img_filename)
                    with open(img_path, 'wb') as f:
                        f.write(img_data)
                    image_map[(row, col)] = img_filename
                    print(f'Extracted image for row={row}, col={col}: {img_filename} ({len(img_data)} bytes)')
                else:
                    image_map[(row, col)] = True  # Image exists but couldn't extract data
                    print(f'Image detected at row={row}, col={col} but could not extract data (size={len(img_data) if img_data else 0})')
            except Exception as e:
                print(f'Image extraction error for img {img_idx}: {e}')

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        row_data = {}
        for col_idx, cell in enumerate(row):
            if col_idx < len(headers):
                val = cell.value if cell.value is not None else ''
                row_data[headers[col_idx]] = str(val).strip()

                # Check for hyperlinks
                if cell.hyperlink and cell.hyperlink.target:
                    hyperlinks[(row_idx, col_idx)] = cell.hyperlink.target
                    row_data[headers[col_idx] + '__hyperlink'] = cell.hyperlink.target

                # Check if this cell has an extracted embedded image
                img_file = image_map.get((row_idx, col_idx))
                if img_file and isinstance(img_file, str):
                    row_data[headers[col_idx] + '__embedded_image'] = img_file

        data.append(row_data)

    df = pd.DataFrame(data)
    return df


def parse_pdf(filepath):
    try:
        import pdfplumber
        tables = []
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                page_tables = page.extract_tables()
                for table in page_tables:
                    if table and len(table) > 1:
                        headers = [str(h or '').strip() for h in table[0]]
                        for row in table[1:]:
                            row_dict = {}
                            for i, val in enumerate(row):
                                if i < len(headers):
                                    row_dict[headers[i]] = str(val or '').strip()
                            tables.append(row_dict)
        if tables:
            return pd.DataFrame(tables)
    except ImportError:
        pass
    raise ValueError('Could not parse PDF. Install pdfplumber: pip install pdfplumber')


# ── Product Lookup Services ──────────────────────────────────────────────────
# Priority order for retail sources (1st tier)
RETAIL_PRIORITY = ['amazon', 'walmart', 'target', 'costco', 'kroger', 'walgreens', 'cvs', 'ebay', 'bestbuy', 'homedepot']

# Browser-like headers for web scraping
BROWSER_HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.5',
}

# Simple in-memory cache to avoid burning rate limits on repeat lookups
_lookup_cache = {}
# Track which APIs are rate-limited — stores timestamp when rate limit was hit
# so we can reset after a cooldown period
_api_rate_limited = {}
_RATE_LIMIT_COOLDOWN = 600  # 10 minutes before retrying a rate-limited API

import time

def _is_rate_limited(api_name):
    """Check if an API is currently rate-limited, with automatic cooldown reset."""
    if api_name not in _api_rate_limited:
        return False
    hit_time = _api_rate_limited[api_name]
    if time.time() - hit_time > _RATE_LIMIT_COOLDOWN:
        print(f'{api_name} rate limit cooldown expired — retrying')
        del _api_rate_limited[api_name]
        return False
    return True

def _set_rate_limited(api_name):
    """Mark an API as rate-limited with current timestamp."""
    _api_rate_limited[api_name] = time.time()
    print(f'{api_name} rate limited — will retry after {_RATE_LIMIT_COOLDOWN}s cooldown')


def _resolve_redirect_url(url):
    """Follow redirect URLs (like UPCitemdb's /noredir/) to get the actual destination.
    Returns the final URL after redirects, or the original URL if resolution fails."""
    if not url:
        return url
    # UPCitemdb redirect pattern: contains upcitemdb.com and a redirect param
    # Also handle any generic redirect/tracking URLs
    is_redirect = (
        'upcitemdb.com' in url or
        '/redirect?' in url or
        '/noredir/' in url or
        'go.redirectingat.com' in url
    )
    if not is_redirect:
        return url
    try:
        # Use HEAD request with redirects to find final URL
        resp = requests.head(url, headers=BROWSER_HEADERS, timeout=5, allow_redirects=True)
        final_url = resp.url
        if final_url and final_url != url:
            print(f'Resolved redirect: {url[:80]} → {final_url[:80]}')
            return final_url
    except Exception as e:
        print(f'Redirect resolution failed for {url[:80]}: {e}')
    # Fallback: try to extract destination from URL params
    try:
        parsed = urllib.parse.urlparse(url)
        params = urllib.parse.parse_qs(parsed.query)
        for key in ('to', 'url', 'dest', 'redirect', 'target', 'u'):
            if key in params:
                dest = params[key][0]
                if dest.startswith('http'):
                    return dest
    except Exception:
        pass
    return url


def _is_product_page_url(url):
    """Check if a URL is an actual product page (not a search/listing page).
    Returns (clean_url, source) or (None, None)."""
    if not url:
        return None, None

    # Amazon: must have /dp/ASIN — reject /s? (search), /b/ (browse), /gp/ (generic)
    amazon_match = re.search(r'amazon\.com.*/dp/([A-Z0-9]{10})', url)
    if amazon_match:
        asin = amazon_match.group(1)
        return f'https://www.amazon.com/dp/{asin}', 'Amazon'

    # Walmart: must have /ip/ (item page)
    walmart_match = re.search(r'(https?://[^/]*walmart\.com/ip/[^\s?#]+)', url)
    if walmart_match:
        return walmart_match.group(1), 'Walmart'

    # Target: must have /-/A- (product ID)
    target_match = re.search(r'(https?://[^/]*target\.com/[^\s?#]*/-/A-\d+)', url)
    if target_match:
        return target_match.group(1), 'Target'

    # Home Depot: must have /p/ (product page)
    hd_match = re.search(r'(https?://[^/]*homedepot\.com/p/[^\s?#]+)', url)
    if hd_match:
        return hd_match.group(1), 'Home Depot'

    # Costco: must have /product (product page)
    costco_match = re.search(r'(https?://[^/]*costco\.com[^\s?#]*/product[^\s?#]*)', url)
    if costco_match:
        return costco_match.group(1), 'Costco'

    # Best Buy: must have /site/ with .p?skuId
    bb_match = re.search(r'(https?://[^/]*bestbuy\.com/site/[^\s?#]+\.p)', url)
    if bb_match:
        return bb_match.group(1), 'Best Buy'

    # Kroger: product page
    kroger_match = re.search(r'(https?://[^/]*kroger\.com/p/[^\s?#]+)', url)
    if kroger_match:
        return kroger_match.group(1), 'Kroger'

    # Walgreens: product page
    wg_match = re.search(r'(https?://[^/]*walgreens\.com/store/id=[^\s?#]+)', url)
    if wg_match:
        return wg_match.group(1), 'Walgreens'

    # CVS: product page
    cvs_match = re.search(r'(https?://[^/]*cvs\.com/shop/[^\s?#]+)', url)
    if cvs_match:
        return cvs_match.group(1), 'CVS'

    # eBay: must have /itm/ (item page)
    ebay_match = re.search(r'(https?://[^/]*ebay\.com/itm/[^\s?#]+)', url)
    if ebay_match:
        return ebay_match.group(1), 'eBay'

    return None, None


# SearXNG public instances — open-source meta search engines that provide
# JSON APIs and are designed for programmatic use (unlike Google/Bing/DDG
# which block datacenter IPs like Render's).
SEARXNG_INSTANCES = [
    'https://search.bus-hit.me',
    'https://searx.be',
    'https://search.sapti.me',
    'https://searx.tiekoetter.com',
    'https://search.ononoki.org',
    'https://searx.oxf.app',
    'https://paulgo.io',
    'https://search.nqdev.ch',
    'https://searx.work',
    'https://search.hbubli.cc',
    'https://searx.foss.family',
    'https://search.mdosch.de',
    'https://searx.namejeff.xyz',
    'https://etsi.me',
]


_searxng_dead_instances = {}  # Track instances that keep failing

def _search_searxng(query, categories='general', max_instances=2):
    """Search using SearXNG public instances (JSON API).
    Tries up to max_instances for speed (default 2).
    Skips instances that have failed recently.
    Returns (url, title, source) or (None, None, None)."""
    import random
    # Filter out recently-dead instances (failed in last 5 min)
    now = time.time()
    alive = [i for i in SEARXNG_INSTANCES if now - _searxng_dead_instances.get(i, 0) > 300]
    if not alive:
        # All dead — reset and try again
        alive = SEARXNG_INSTANCES.copy()
        _searxng_dead_instances.clear()
    random.shuffle(alive)

    tried = 0
    for instance in alive:
        if tried >= max_instances:
            break
        tried += 1
        try:
            resp = requests.get(
                f'{instance}/search',
                params={
                    'q': query,
                    'format': 'json',
                    'categories': categories,
                    'language': 'en',
                },
                headers=BROWSER_HEADERS,
                timeout=6  # Short timeout — fail fast, try next
            )
            if resp.status_code == 200:
                data = resp.json()
                results = data.get('results', [])
                for r in results:
                    url = r.get('url', '')
                    title = r.get('title', '')
                    if not url:
                        continue
                    clean_url, source = _is_product_page_url(url)
                    if clean_url and source:
                        # Clean up title
                        title = re.sub(r'^(Amazon\.com|Walmart\.com|Target|Home Depot|Costco)\s*[:\-–]\s*', '', title).strip()
                        title = re.sub(r'\s*[:\-–|]\s*(Amazon\.com|Walmart|Target|Home Depot|Costco|Best Buy|eBay)\s*$', '', title).strip()
                        print(f'SearXNG ({instance}) found: {clean_url} — "{title}"')
                        return clean_url, title, source
                # Had results but none were product pages — don't try more instances,
                # the query itself probably won't match on other instances either
                print(f'SearXNG ({instance}) returned {len(results)} results but no product pages for: {query}')
                break
            else:
                print(f'SearXNG ({instance}) returned status {resp.status_code}')
                _searxng_dead_instances[instance] = time.time()
        except Exception as e:
            print(f'SearXNG ({instance}) error: {e}')
            _searxng_dead_instances[instance] = time.time()
            continue

    return None, None, None


def _search_retailer_directly(upc, name=None):
    """Search major retailer websites directly by UPC/name and scrape product links.
    This bypasses search engine limitations by going straight to retailer search pages.
    IMPORTANT: Detects "no results" pages to avoid returning random suggested products.
    Returns (url, title, source) or (None, None, None)."""

    upc_clean = re.sub(r'[^0-9]', '', str(upc)) if upc else ''
    search_term = upc_clean if upc_clean else str(name or '').strip()
    if not search_term:
        return None, None, None

    # Retailer search URLs — these are the internal search pages
    retailer_searches = [
        {
            'name': 'Walmart',
            'url': f'https://www.walmart.com/search?q={urllib.parse.quote_plus(search_term)}',
            'product_pattern': r'walmart\.com/ip/',
            # Walmart "no results" indicators
            'no_results_patterns': [r'no results for', r'0 results for', r'try searching for something else'],
        },
        {
            'name': 'Amazon',
            'url': f'https://www.amazon.com/s?k={urllib.parse.quote_plus(search_term)}',
            'product_pattern': r'amazon\.com.*/dp/[A-Z0-9]{10}',
            # Amazon "no results" — shows "No results for" then suggests random products
            'no_results_patterns': [r'No results for', r'did not match any products', r'SUGGESTED', r'Consider these alternative'],
        },
        {
            'name': 'Target',
            'url': f'https://www.target.com/s?searchTerm={urllib.parse.quote_plus(search_term)}',
            'product_pattern': r'target\.com/.*/-/A-\d+',
            'no_results_patterns': [r'No results found', r'we couldn\'t find', r'try a new search'],
        },
        {
            'name': 'Home Depot',
            'url': f'https://www.homedepot.com/s/{urllib.parse.quote_plus(search_term)}',
            'product_pattern': r'homedepot\.com/p/',
            'no_results_patterns': [r'no results', r'0 results', r'did not match'],
        },
        {
            'name': 'Best Buy',
            'url': f'https://www.bestbuy.com/site/searchpage.jsp?st={urllib.parse.quote_plus(search_term)}',
            'product_pattern': r'bestbuy\.com/site/.*\.p',
            'no_results_patterns': [r'No results found', r'0 Results', r'did not return any results'],
        },
    ]

    for retailer in retailer_searches:
        try:
            resp = requests.get(
                retailer['url'],
                headers=BROWSER_HEADERS,
                timeout=6,
                allow_redirects=True
            )
            if resp.status_code != 200:
                print(f"Direct {retailer['name']} search returned {resp.status_code}")
                continue

            page_text = resp.text

            # ── CRITICAL: Check for "no results" page BEFORE scraping links ──
            # Retailers show random suggested/sponsored products when UPC has no match.
            # We must detect this and skip — otherwise we return wrong products.
            is_no_results = False
            for pattern in retailer.get('no_results_patterns', []):
                if re.search(pattern, page_text, re.IGNORECASE):
                    print(f"Direct {retailer['name']} search: NO RESULTS page detected for '{search_term}'")
                    is_no_results = True
                    break
            if is_no_results:
                continue

            # Check if we were redirected directly to a product page
            final_url = resp.url
            clean_url, source = _is_product_page_url(final_url)
            if clean_url:
                print(f"Direct {retailer['name']} search redirected to product: {clean_url}")
                return clean_url, None, source

            # Parse the search results page for product links
            soup = BeautifulSoup(page_text, 'html.parser')
            product_links = []

            # Find all links matching product page patterns
            seen_urls = set()
            for a in soup.find_all('a', href=True):
                href = a['href']
                # Make relative URLs absolute
                if href.startswith('/'):
                    parsed_url = urllib.parse.urlparse(retailer['url'])
                    href = f'{parsed_url.scheme}://{parsed_url.netloc}{href}'
                if re.search(retailer['product_pattern'], href):
                    clean_url, source = _is_product_page_url(href)
                    if clean_url and clean_url not in seen_urls:
                        seen_urls.add(clean_url)
                        product_links.append((clean_url, source))

            if product_links:
                # Return the first product link found (top search result)
                url, source = product_links[0]
                print(f"Direct {retailer['name']} search found product: {url}")
                return url, None, source

            print(f"Direct {retailer['name']} search: no product links found in results")
        except Exception as e:
            print(f"Direct {retailer['name']} search error: {e}")
            continue

    return None, None, None


def _fetch_product_title_from_page(url):
    """Fetch a product page and extract the title. Returns title string or None."""
    try:
        resp = requests.get(url, headers=BROWSER_HEADERS, timeout=6, allow_redirects=True)
        if resp.status_code == 200:
            soup = BeautifulSoup(resp.text, 'html.parser')

            # Amazon product title
            if 'amazon.com' in url:
                el = soup.find('span', id='productTitle')
                if el:
                    return el.get_text(strip=True)

            # Walmart product title
            if 'walmart.com' in url:
                el = soup.find('h1', itemprop='name') or soup.find('h1')
                if el:
                    return el.get_text(strip=True)

            # Target product title
            if 'target.com' in url:
                el = soup.find('h1') or soup.find('title')
                if el:
                    t = el.get_text(strip=True)
                    return re.sub(r'\s*[:\-–|]\s*Target$', '', t)

            # Home Depot product title
            if 'homedepot.com' in url:
                el = soup.find('h1', class_='product-details__title') or soup.find('h1')
                if el:
                    return el.get_text(strip=True)

            # Generic: use <title> tag, clean up common suffixes
            title_tag = soup.find('title')
            if title_tag:
                t = title_tag.get_text(strip=True)
                t = re.sub(r'\s*[:\-–|]\s*(Amazon\.com|Walmart|Target|Home Depot|Costco|Best Buy|eBay).*$', '', t)
                return t.strip()
    except Exception as e:
        print(f'Page fetch error for {url}: {e}')
    return None


def _fetch_product_image_from_page(url):
    """Fetch a product page and extract the main product image URL.
    Returns image URL string or None."""
    try:
        resp = requests.get(url, headers=BROWSER_HEADERS, timeout=6, allow_redirects=True)
        if resp.status_code != 200:
            return None
        soup = BeautifulSoup(resp.text, 'html.parser')

        # Amazon: main product image
        if 'amazon.com' in url:
            el = soup.find('img', id='landingImage')
            if el and el.get('src', '').startswith('http'):
                return el['src']
            # Fallback: look for data-old-hires attribute
            if el and el.get('data-old-hires', '').startswith('http'):
                return el['data-old-hires']

        # Walmart: product image
        if 'walmart.com' in url:
            el = soup.find('img', {'data-testid': 'hero-image'})
            if not el:
                el = soup.find('img', itemprop='image')
            if el and el.get('src', '').startswith('http'):
                return el['src']

        # Target: product image
        if 'target.com' in url:
            el = soup.find('img', {'data-test': 'product-image'})
            if el and el.get('src', '').startswith('http'):
                return el['src']

        # Home Depot: product image
        if 'homedepot.com' in url:
            el = soup.find('img', class_='stretchy')
            if not el:
                el = soup.find('img', id='mainImage')
            if el and el.get('src', '').startswith('http'):
                return el['src']

        # Generic: try og:image meta tag (most retail sites have this)
        og_img = soup.find('meta', property='og:image')
        if og_img and og_img.get('content', '').startswith('http'):
            return og_img['content']

        # Fallback: try itemprop="image"
        item_img = soup.find('img', itemprop='image')
        if item_img and item_img.get('src', '').startswith('http'):
            return item_img['src']
        # itemprop on a link/meta
        item_meta = soup.find('meta', itemprop='image')
        if item_meta and item_meta.get('content', '').startswith('http'):
            return item_meta['content']

    except Exception as e:
        print(f'Image scrape error for {url}: {e}')
    return None


def _name_similarity(name1, name2):
    """Smart word-overlap similarity between two product names. Returns 0.0 to 1.0.
    Handles cases where original is short ('Tide Pods') and lookup is long
    ('Tide PODS Laundry Detergent Soap Pacs, 81 Count, Spring Meadow').
    The key question: do most words from the SHORT name appear in the LONG name?"""
    if not name1 or not name2:
        return 0.0
    # Normalize: lowercase, remove punctuation, split into words
    words1 = set(re.sub(r'[^a-z0-9\s]', '', name1.lower()).split())
    words2 = set(re.sub(r'[^a-z0-9\s]', '', name2.lower()).split())
    # Remove very common filler words
    stop = {'the', 'a', 'an', 'and', 'or', 'of', 'for', 'in', 'with', 'to', 'by',
            'is', 'at', 'on', 'from', 'pack', 'ct', 'oz', 'lb', 'count', 'ea',
            'each', 'per', 'size', 'new', 'free', 'buy', 'best', 'top', 'item'}
    words1 = words1 - stop
    words2 = words2 - stop
    if not words1 or not words2:
        return 0.0
    overlap = words1 & words2

    # Use the SHORTER name as the reference — "do most words from the original
    # name appear somewhere in the lookup result?"
    shorter = words1 if len(words1) <= len(words2) else words2
    if not shorter:
        return 0.0
    return len(overlap) / len(shorter)


def _verify_upc_on_page(url, upc_clean, page_text=None):
    """Check if a UPC number actually appears on a product page.
    If the UPC doesn't appear, the page is probably a random/suggested product,
    not a genuine match for this barcode. Returns True if verified."""
    if not upc_clean or len(upc_clean) < 8:
        return True  # Can't verify short/invalid UPCs, assume OK
    try:
        if page_text is None:
            resp = requests.get(url, headers=BROWSER_HEADERS, timeout=6, allow_redirects=True)
            if resp.status_code != 200:
                return False
            page_text = resp.text
        # Check if the UPC appears anywhere on the page
        if upc_clean in page_text:
            return True
        # Also check with common formatting (dashes, spaces)
        if len(upc_clean) == 12:
            formatted = f'{upc_clean[:1]}-{upc_clean[1:6]}-{upc_clean[6:11]}-{upc_clean[11:]}'
            if formatted in page_text:
                return True
        print(f'UPC {upc_clean} NOT found on page {url[:60]} — likely wrong product')
        return False
    except Exception as e:
        print(f'UPC page verification error: {e}')
        return False


def search_product_on_web(upc=None, name=None):
    """Search for a product on major retail sites. Priority order:
    1. Direct retailer search (fastest, most reliable from datacenter IPs)
    2. SearXNG meta-search (fallback only — many instances block datacenter IPs)
    ONLY returns results that are actual product pages.
    Returns dict with keys: url, title, source."""
    result = {'url': None, 'title': None, 'source': None}

    upc_clean = re.sub(r'[^0-9]', '', str(upc)) if upc else ''
    has_upc = bool(upc_clean) and upc_clean not in ('0', '')

    def _validate_and_return(url, source, original_name, searched_by_upc=False):
        """Fetch page title and validate the result is the correct product.
        Three checks:
        1. If we searched by UPC, verify the UPC appears on the product page
        2. If we can't fetch a page title, reject (can't verify)
        3. Name similarity must pass threshold (higher for short names)
        Returns result dict or None."""
        try:
            # Fetch the actual product page
            resp = requests.get(url, headers=BROWSER_HEADERS, timeout=6, allow_redirects=True)
            if resp.status_code != 200:
                print(f'Page fetch failed ({resp.status_code}) for {url[:60]}')
                return None
            page_text = resp.text
        except Exception as e:
            print(f'Page fetch error for {url[:60]}: {e}')
            return None

        # CHECK 1: If searched by UPC, verify UPC appears on the page
        if searched_by_upc and has_upc:
            if not _verify_upc_on_page(url, upc_clean, page_text=page_text):
                return None

        # Extract page title
        page_title = None
        try:
            soup = BeautifulSoup(page_text, 'html.parser')
            # Amazon
            if 'amazon.com' in url:
                el = soup.find('span', id='productTitle')
                if el:
                    page_title = el.get_text(strip=True)
            # Walmart
            elif 'walmart.com' in url:
                el = soup.find('h1', itemprop='name') or soup.find('h1')
                if el:
                    page_title = el.get_text(strip=True)
            # Target
            elif 'target.com' in url:
                el = soup.find('h1') or soup.find('title')
                if el:
                    t = el.get_text(strip=True)
                    page_title = re.sub(r'\s*[:\-–|]\s*Target$', '', t)
            # Home Depot
            elif 'homedepot.com' in url:
                el = soup.find('h1', class_='product-details__title') or soup.find('h1')
                if el:
                    page_title = el.get_text(strip=True)
            # Generic fallback
            if not page_title:
                title_tag = soup.find('title')
                if title_tag:
                    t = title_tag.get_text(strip=True)
                    page_title = re.sub(r'\s*[:\-–|]\s*(Amazon\.com|Walmart|Target|Home Depot|Costco|Best Buy|eBay).*$', '', t).strip()
        except Exception:
            pass

        # CHECK 2: If we can't get a title, reject — we can't verify this is right
        if not page_title:
            print(f'Could not extract title from {url[:60]} — rejecting')
            return None

        # CHECK 3: Name similarity — require higher threshold for short names
        if original_name:
            sim = _name_similarity(original_name, page_title)
            # Short names (1-3 meaningful words) need higher threshold because
            # a single shared category word (e.g., "light") can cause false matches
            clean_words = set(re.sub(r'[^a-z0-9\s]', '', original_name.lower()).split())
            stop = {'the', 'a', 'an', 'and', 'or', 'of', 'for', 'in', 'with', 'to', 'by',
                    'is', 'at', 'on', 'from', 'pack', 'ct', 'oz', 'lb', 'count', 'ea',
                    'each', 'per', 'size', 'new', 'free', 'buy', 'best', 'top', 'item'}
            meaningful_words = clean_words - stop
            min_sim = 0.55 if len(meaningful_words) <= 3 else 0.35

            if sim < min_sim:
                print(f'Result rejected (sim={sim:.2f} < {min_sim}): "{page_title[:60]}" vs "{original_name[:60]}"')
                return None
            print(f'Result accepted (sim={sim:.2f} >= {min_sim}): "{page_title[:60]}" for "{original_name[:60]}"')

        return {'url': url, 'title': page_title, 'source': source}

    # ── Strategy 1: Direct retailer search (FAST — goes straight to retailer sites) ──
    # Search by UPC first, then by name
    if has_upc:
        try:
            direct_url, _, direct_source = _search_retailer_directly(upc, None)
            if direct_url:
                r = _validate_and_return(direct_url, direct_source, name, searched_by_upc=True)
                if r:
                    print(f'Direct UPC retailer search hit: {direct_url}')
                    return r
        except Exception as e:
            print(f'Direct UPC retailer search failed: {e}')

    # Direct search by product name
    if name:
        try:
            direct_url, _, direct_source = _search_retailer_directly(None, name)
            if direct_url:
                r = _validate_and_return(direct_url, direct_source, name, searched_by_upc=False)
                if r:
                    print(f'Direct name retailer search hit: {direct_url}')
                    return r
        except Exception as e:
            print(f'Direct name retailer search failed: {e}')

    # ── Strategy 2: SearXNG meta-search (SLOW fallback — most instances fail from datacenter) ──
    if has_upc:
        try:
            url, snippet_title, source = _search_searxng(f'{upc_clean}', max_instances=2)
            if url:
                r = _validate_and_return(url, source, name, searched_by_upc=True)
                if r:
                    return r
        except Exception:
            pass

    if name:
        try:
            clean_name = str(name).strip()
            url, snippet_title, source = _search_searxng(f'{clean_name} buy', max_instances=2)
            if url:
                r = _validate_and_return(url, source, name, searched_by_upc=False)
                if r:
                    return r
        except Exception:
            pass

    return result


def lookup_product_info(upc=None, name=None):
    """Unified product lookup: returns dict with title, image, retail_link, source, upc_matched.
    upc_matched=True means the result came from a direct UPC barcode lookup (trusted).
    upc_matched=False means it came from a web search (needs validation)."""
    result = {'title': None, 'image': None, 'retail_link': None, 'source': None, 'upc_matched': False}

    upc_clean = re.sub(r'[^0-9]', '', str(upc)) if upc else ''
    has_upc = bool(upc_clean) and upc_clean.lower() not in ('na', 'nan', 'none', '0')

    # Check cache first
    cache_key = f'{upc_clean}|{name or ""}'
    if cache_key in _lookup_cache:
        print(f'Cache hit for: {cache_key}')
        return _lookup_cache[cache_key].copy()

    # ── 1. UPCitemdb (best source — returns title, images, and offers in one call) ──
    if has_upc and not _is_rate_limited('upcitemdb'):
        try:
            resp = requests.get(
                f'https://api.upcitemdb.com/prod/trial/lookup?upc={upc_clean}',
                timeout=5,
                headers={'Accept': 'application/json'}
            )
            if resp.status_code == 429:
                print('UPCitemdb rate limited — skipping for remaining items')
                _set_rate_limited('upcitemdb')
            elif resp.status_code == 200:
                data = resp.json()
                items = data.get('items', [])
                if items:
                    item = items[0]
                    # UPC API matched — this is a direct barcode lookup, trusted
                    result['upc_matched'] = True
                    # Get product title
                    title = item.get('title', '').strip()
                    if title:
                        result['title'] = title

                    # Get product image
                    images = item.get('images', [])
                    if images:
                        result['image'] = images[0]

                    # Get retail link — prioritize Amazon, then other 1st-tier retailers
                    # UPCitemdb offers are from a direct UPC match, so they're trustworthy.
                    # Just clean up URLs where possible and skip obvious search pages.
                    offers = item.get('offers', [])
                    best_link = None
                    best_priority = len(RETAIL_PRIORITY) + 1  # worst priority
                    fallback_link = None

                    for offer in offers:
                        link = offer.get('link', '')
                        merchant = offer.get('merchant', '').lower()
                        if not link:
                            continue

                        # Resolve redirect URLs (UPCitemdb uses redirect/tracking URLs)
                        link = _resolve_redirect_url(link)

                        # Try to clean up to a canonical product URL
                        product_url, url_source = _is_product_page_url(link)
                        if product_url:
                            link = product_url

                        # Check against priority list
                        for idx, retailer in enumerate(RETAIL_PRIORITY):
                            if retailer in merchant or (url_source and retailer in url_source.lower()):
                                if idx < best_priority:
                                    best_priority = idx
                                    best_link = link
                                    result['source'] = url_source or retailer.capitalize()
                                break
                        else:
                            # Not in priority list — keep as fallback only if it's a real retailer URL
                            if not fallback_link and not 'upcitemdb.com' in link:
                                fallback_link = link

                    result['retail_link'] = best_link or fallback_link

                    # If we got everything, return early
                    if result['title'] and result['image'] and result['retail_link']:
                        _lookup_cache[cache_key] = result.copy()
                        return result
        except Exception as e:
            print(f'UPCitemdb lookup error: {e}')

    # ── 2. Open Food Facts (good for food/grocery — has title + image) ──
    if has_upc and (not result['title'] or not result['image']):
        try:
            resp = requests.get(
                f'https://world.openfoodfacts.org/api/v0/product/{upc_clean}.json',
                timeout=5
            )
            if resp.status_code == 200:
                data = resp.json()
                product = data.get('product', {})
                if product:
                    result['upc_matched'] = True  # Direct UPC lookup
                    if not result['title']:
                        off_name = product.get('product_name', '').strip()
                        if off_name:
                            result['title'] = off_name
                            if not result['source']:
                                result['source'] = 'Open Food Facts'
                    if not result['image']:
                        img = (product.get('image_url') or
                               product.get('image_front_url') or
                               product.get('image_front_small_url'))
                        if img:
                            result['image'] = img
        except Exception:
            pass

    # ── 3. Open Food Facts search by name (fallback for food/grocery items) ──
    if not result['image'] and name:
        try:
            q = urllib.parse.quote_plus(str(name))
            resp = requests.get(
                f'https://world.openfoodfacts.org/cgi/search.pl?search_terms={q}&search_simple=1&action=process&json=1&page_size=1',
                timeout=5
            )
            if resp.status_code == 200:
                data = resp.json()
                products = data.get('products', [])
                if products:
                    if not result['title']:
                        off_name = products[0].get('product_name', '').strip()
                        if off_name:
                            result['title'] = off_name
                    img = (products[0].get('image_url') or
                           products[0].get('image_front_url') or
                           products[0].get('image_front_small_url'))
                    if img:
                        result['image'] = img
        except Exception:
            pass

    # ── 6. Web search fallback: find actual product page (not just search URL) ──
    needs_link = not result['retail_link']
    needs_title = not result['title']
    if needs_link or needs_title:
        web_result = search_product_on_web(upc=upc, name=name)
        if web_result['url'] and needs_link:
            result['retail_link'] = web_result['url']
        if web_result['title'] and needs_title:
            result['title'] = web_result['title']
        if web_result['source'] and not result['source']:
            result['source'] = web_result['source']

    # ── 7. No fallback to search URLs ──
    # We intentionally do NOT generate amazon.com/s?k= search URLs as fallback.
    # It's better to leave the link blank than to give a search page that may
    # show unrelated products. The enrichment UI will show a dash for missing links.

    # Cache the result (even if empty, to avoid re-hitting rate-limited APIs)
    _lookup_cache[cache_key] = result.copy()

    return result


# ── Image Download Helper ────────────────────────────────────────────────────
def download_image(url, max_size_kb=500):
    """Download an image from URL and return as bytes. Returns (bytes, format) or (None, None)."""
    if not url or not url.startswith('http'):
        return None, None
    # Skip google search URLs
    if 'google.com/search' in url:
        return None, None
    try:
        resp = requests.get(url, timeout=6, stream=True, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
        if resp.status_code != 200:
            return None, None
        content_type = resp.headers.get('Content-Type', '').lower()
        if 'image' not in content_type and 'octet-stream' not in content_type:
            return None, None

        img_data = resp.content
        if len(img_data) > max_size_kb * 1024 * 5:  # max ~2.5MB
            return None, None

        # Determine format
        if 'png' in content_type or url.lower().endswith('.png'):
            fmt = 'png'
        elif 'gif' in content_type or url.lower().endswith('.gif'):
            fmt = 'gif'
        elif 'webp' in content_type or url.lower().endswith('.webp'):
            fmt = 'png'  # will need conversion
        else:
            fmt = 'jpeg'

        return img_data, fmt
    except Exception:
        return None, None


def resize_image_bytes(img_bytes, max_width=120, max_height=120):
    """Resize image bytes to fit within max dimensions. Returns (bytes, format)."""
    try:
        from PIL import Image as PILImage
        img = PILImage.open(io.BytesIO(img_bytes))
        # Convert to RGB if needed
        if img.mode in ('RGBA', 'P', 'LA'):
            background = PILImage.new('RGB', img.size, (255, 255, 255))
            if img.mode == 'P':
                img = img.convert('RGBA')
            background.paste(img, mask=img.split()[-1] if 'A' in img.mode else None)
            img = background
        elif img.mode != 'RGB':
            img = img.convert('RGB')

        # Resize
        img.thumbnail((max_width, max_height), PILImage.LANCZOS)

        buf = io.BytesIO()
        img.save(buf, format='PNG', optimize=True)
        buf.seek(0)
        return buf.getvalue(), 'png'
    except ImportError:
        # Pillow not installed, return original
        return img_bytes, 'jpeg'
    except Exception:
        return img_bytes, 'jpeg'


# ── Column Auto-Mapping ──────────────────────────────────────────────────────
COLUMN_ALIASES = {
    'Image': ['image', 'img', 'photo', 'picture', 'product image', 'thumbnail', 'image url', 'image link', 'product photo'],
    'Item Name': ['item name', 'product name', 'name', 'description', 'item description', 'product',
                  'item', 'title', 'product description', 'product title'],
    'Expiration': ['expiration', 'exp', 'exp date', 'expiry', 'expiry date', 'best by', 'use by', 'bb date', 'sell by'],
    'UPC/Item #': ['upc', 'upc code', 'item #', 'item number', 'sku', 'barcode', 'gtin', 'ean',
                   'item#', 'upc/item #', 'product code', 'asin'],
    'Quantity': ['quantity', 'qty', 'units', 'count', 'amount', 'pcs', 'pieces', 'total qty', 'available'],
    'Casepack': ['casepack', 'case pack', 'case qty', 'pack size', 'inner pack', 'units per case', 'per case', 'case'],
    'Cost': ['cost', 'price', 'unit cost', 'unit price', 'wholesale', 'wholesale price', 'our price',
             'your cost', 'net price', 'each'],
    'Retail Link': ['retail link', 'link', 'url', 'product link', 'product url', 'retail url', 'buy link',
                    'store link', 'website'],
    'FOB': ['fob', 'f.o.b.', 'fob location', 'ship from', 'origin', 'warehouse', 'location', 'ship point', 'freight'],
}


def auto_map_columns(source_columns):
    mapping = {}
    source_lower = {c: c.lower().strip() for c in source_columns}
    for template_col, aliases in COLUMN_ALIASES.items():
        for src_col, src_lower in source_lower.items():
            if src_lower in aliases or src_lower == template_col.lower():
                mapping[template_col] = src_col
                break
    return mapping


# ── Routes ────────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html')


@app.route('/serve_image/<filename>')
def serve_image(filename):
    """Serve extracted images from uploaded files."""
    safe = secure_filename(filename)
    path = os.path.join(UPLOAD_FOLDER, safe)
    if os.path.exists(path):
        return send_file(path, mimetype='image/png')
    return '', 404


@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    if file.filename == '' or not allowed_file(file.filename):
        return jsonify({'error': 'Invalid file. Upload .xlsx, .csv, .tsv, or .pdf'}), 400

    filename = secure_filename(file.filename)
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    file.save(filepath)

    try:
        df = parse_upload(filepath, filename)
        session_id = os.urandom(8).hex()
        cache_path = os.path.join(UPLOAD_FOLDER, f'{session_id}.json')
        df.to_json(cache_path, orient='records')

        source_columns = list(df.columns)
        visible_columns = [c for c in source_columns if not c.endswith('__hyperlink') and not c.endswith('__embedded_image')]
        suggested_mapping = auto_map_columns(visible_columns)
        preview = df.head(5).to_dict(orient='records')

        return jsonify({
            'session_id': session_id,
            'source_columns': visible_columns,
            'suggested_mapping': suggested_mapping,
            'preview': preview,
            'row_count': len(df)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/process', methods=['POST'])
def process_data():
    data = request.json
    session_id = data.get('session_id')
    mapping = data.get('mapping', {})

    cache_path = os.path.join(UPLOAD_FOLDER, f'{session_id}.json')
    if not os.path.exists(cache_path):
        return jsonify({'error': 'Session expired. Please re-upload.'}), 400

    df = pd.read_json(cache_path, dtype=str).fillna('')
    results = []

    for idx, row in df.iterrows():
        item = {}
        for template_col in TEMPLATE_COLUMNS:
            src_col = mapping.get(template_col, '')
            if src_col and src_col in df.columns:
                item[template_col] = str(row.get(src_col, '')).strip()
            else:
                item[template_col] = ''

        # Handle expiration
        if not item.get('Expiration') or item['Expiration'].lower() in ('', 'nan', 'none', 'null'):
            item['Expiration'] = 'NA'

        # Check for hyperlinks in source (Retail Link)
        src_link_col = mapping.get('Retail Link', '')
        if src_link_col:
            hyperlink_key = src_link_col + '__hyperlink'
            if hyperlink_key in df.columns and row.get(hyperlink_key):
                item['Retail Link'] = row[hyperlink_key]

        # Check for image hyperlinks
        src_img_col = mapping.get('Image', '')
        if src_img_col:
            hyperlink_key = src_img_col + '__hyperlink'
            if hyperlink_key in df.columns and row.get(hyperlink_key):
                item['Image'] = row[hyperlink_key]

            # Check for embedded images extracted from the Excel file
            embedded_key = src_img_col + '__embedded_image'
            if embedded_key in df.columns and row.get(embedded_key):
                # Serve the extracted image via our Flask route
                item['Image'] = '/serve_image/' + row[embedded_key]

        # Also check ALL columns for embedded images if Image is still empty
        if not item.get('Image') or not item['Image'].startswith(('http', '/')):
            for col_name in df.columns:
                if col_name.endswith('__embedded_image') and row.get(col_name):
                    item['Image'] = '/serve_image/' + row[col_name]
                    break

        results.append(item)

    # Save processed results
    result_path = os.path.join(UPLOAD_FOLDER, f'{session_id}_processed.json')
    with open(result_path, 'w') as f:
        json.dump(results, f)

    return jsonify({'results': results, 'session_id': session_id})


@app.route('/debug_lookup')
def debug_lookup():
    """Temporary debug endpoint to diagnose API lookup failures."""
    import traceback
    upc = request.args.get('upc', '805106960230')
    name = request.args.get('name', 'Mini Light')
    results = {'upc': upc, 'name': name, 'tests': {}}

    upc_clean = re.sub(r'[^0-9]', '', str(upc))

    # Test 1: UPCitemdb
    try:
        resp = requests.get(
            f'https://api.upcitemdb.com/prod/trial/lookup?upc={upc_clean}',
            timeout=6, headers={'Accept': 'application/json'}
        )
        results['tests']['upcitemdb'] = {
            'status_code': resp.status_code,
            'body_preview': resp.text[:500],
            'items_count': len(resp.json().get('items', [])) if resp.status_code == 200 else None
        }
    except Exception as e:
        results['tests']['upcitemdb'] = {'error': str(e), 'traceback': traceback.format_exc()}

    # Test 2: Open Food Facts
    try:
        resp = requests.get(
            f'https://world.openfoodfacts.org/api/v0/product/{upc_clean}.json',
            timeout=6
        )
        results['tests']['openfoodfacts'] = {
            'status_code': resp.status_code,
            'status_field': resp.json().get('status') if resp.status_code == 200 else None
        }
    except Exception as e:
        results['tests']['openfoodfacts'] = {'error': str(e)}

    # Test 3: SearXNG (replaces DuckDuckGo/Google/Bing which block datacenter IPs)
    try:
        q = f'{upc_clean} buy'
        url, snippet_title, source = _search_searxng(q)
        page_title = _fetch_product_title_from_page(url) if url else None
        results['tests']['searxng'] = {
            'query': q,
            'found_url': url,
            'snippet_title': snippet_title,
            'page_title': page_title,
            'found_source': source,
        }
    except Exception as e:
        results['tests']['searxng'] = {'error': str(e), 'traceback': traceback.format_exc()}

    # Test 3b: SearXNG with name query
    try:
        q = f'{name} buy amazon'
        url2, snippet_title2, source2 = _search_searxng(q)
        page_title2 = _fetch_product_title_from_page(url2) if url2 else None
        results['tests']['searxng_name'] = {
            'query': q,
            'found_url': url2,
            'snippet_title': snippet_title2,
            'page_title': page_title2,
            'found_source': source2,
        }
    except Exception as e:
        results['tests']['searxng_name'] = {'error': str(e)}

    # Test 3c: SearXNG shopping category
    try:
        q = f'{upc_clean}'
        url3, snippet_title3, source3 = _search_searxng(q, categories='shopping')
        results['tests']['searxng_shopping'] = {
            'query': q,
            'found_url': url3,
            'snippet_title': snippet_title3,
            'found_source': source3,
        }
    except Exception as e:
        results['tests']['searxng_shopping'] = {'error': str(e)}

    # Test 3d: Direct retailer search
    try:
        d_url, d_title, d_source = _search_retailer_directly(upc, name)
        d_page_title = _fetch_product_title_from_page(d_url) if d_url else None
        results['tests']['direct_retailer_search'] = {
            'search_term': upc_clean or name,
            'found_url': d_url,
            'page_title': d_page_title,
            'found_source': d_source,
        }
    except Exception as e:
        results['tests']['direct_retailer_search'] = {'error': str(e), 'traceback': traceback.format_exc()}

    # Test 4: Full lookup_product_info
    try:
        info = lookup_product_info(upc=upc, name=name)
        results['tests']['lookup_product_info'] = info
    except Exception as e:
        results['tests']['lookup_product_info'] = {'error': str(e), 'traceback': traceback.format_exc()}

    return jsonify(results)


def _enrich_single_item(i, items):
    """Enrich a single item — designed to run in a thread pool."""
    if i >= len(items):
        return None
    item = items[i].copy()
    upc = item.get('UPC/Item #', '')
    original_name = item.get('Item Name', '').strip()

    # Check what's missing
    img_val = item.get('Image', '')
    needs_image = not img_val or (not img_val.startswith('http') and not img_val.startswith('/serve_image/'))
    link_val = item.get('Retail Link', '')
    needs_link = not link_val or not link_val.startswith('http')

    # Look up product info
    info = lookup_product_info(upc=upc, name=original_name)

    # ── TRUST LEVEL ──
    if info['upc_matched']:
        lookup_is_relevant = True
        if info['title']:
            print(f'Enrichment UPC MATCH: "{original_name}" → "{info["title"]}" (source: {info["source"]})')
    elif info['title'] and original_name:
        sim = _name_similarity(original_name, info['title'])
        # Higher threshold for short names (1-3 meaningful words) since one
        # shared category word can cause false matches
        clean_words = set(re.sub(r'[^a-z0-9\s]', '', original_name.lower()).split())
        _stop = {'the', 'a', 'an', 'and', 'or', 'of', 'for', 'in', 'with', 'to', 'by',
                'is', 'at', 'on', 'from', 'pack', 'ct', 'oz', 'lb', 'count', 'ea',
                'each', 'per', 'size', 'new', 'free', 'buy', 'best', 'top', 'item'}
        meaningful_words = clean_words - _stop
        min_sim = 0.55 if len(meaningful_words) <= 3 else 0.35
        if sim >= min_sim:
            lookup_is_relevant = True
            print(f'Enrichment NAME MATCH (sim={sim:.2f} >= {min_sim}): "{original_name}" → "{info["title"]}"')
        else:
            lookup_is_relevant = False
            print(f'Enrichment REJECTED (sim={sim:.2f} < {min_sim}): "{original_name}" ≠ "{info["title"]}" — keeping original')
    elif info['title'] and not original_name:
        lookup_is_relevant = True
    else:
        lookup_is_relevant = False

    # ── Item Name enrichment ──
    if lookup_is_relevant and info['title']:
        item['Item Name'] = info['title']
        if info['source']:
            item['_name_source'] = info['source']

    # ── Image enrichment ──
    if needs_image and info['image'] and lookup_is_relevant:
        item['Image'] = info['image']

    # ── Retail link enrichment ──
    if needs_link and info['retail_link'] and lookup_is_relevant:
        retail_url = info['retail_link']
        is_search_page = (
            '/s?' in retail_url or
            '/s/' in retail_url or
            '/search?' in retail_url or
            '/search/' in retail_url or
            'query=' in retail_url or
            'keywords=' in retail_url
        )
        product_url, _ = _is_product_page_url(retail_url)
        if product_url:
            item['Retail Link'] = product_url
        elif not is_search_page:
            item['Retail Link'] = retail_url

    # ── Image from retail link fallback ──
    img_val_now = item.get('Image', '')
    still_needs_image = not img_val_now or (not img_val_now.startswith('http') and not img_val_now.startswith('/serve_image/'))
    retail_link_now = item.get('Retail Link', '')
    if still_needs_image and retail_link_now and retail_link_now.startswith('http'):
        try:
            scraped_img = _fetch_product_image_from_page(retail_link_now)
            if scraped_img:
                item['Image'] = scraped_img
                print(f'Scraped image from retail page: {retail_link_now[:60]} → {scraped_img[:80]}')
        except Exception as e:
            print(f'Image scrape failed for {retail_link_now[:60]}: {e}')

    return {'index': i, 'item': item}


@app.route('/enrich', methods=['POST'])
def enrich_data():
    """Look up images, retail links, and product titles for items that need them.
    Uses thread pool for parallel lookups to dramatically speed up enrichment."""
    from concurrent.futures import ThreadPoolExecutor, as_completed

    data = request.json
    session_id = data.get('session_id')
    items = data.get('items', [])
    indices = data.get('indices', [])

    # Process items in parallel — 3 threads balances speed vs Render free tier resources
    enriched = []
    with ThreadPoolExecutor(max_workers=3) as executor:
        futures = {executor.submit(_enrich_single_item, i, items): i for i in indices}
        for future in as_completed(futures):
            try:
                result = future.result()
                if result:
                    enriched.append(result)
            except Exception as e:
                print(f'Enrichment thread error for index {futures[future]}: {e}')

    # Sort by index to maintain order
    enriched.sort(key=lambda x: x['index'])

    return jsonify({'enriched': enriched})


@app.route('/export', methods=['POST'])
def export_excel():
    data = request.json
    items = data.get('items', [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # Header styling matching DLX template
    header_font_white = Font(name='Aptos Narrow', size=16, bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='4472C4')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_font = Font(name='Aptos Narrow', size=11)
    cell_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    link_font = Font(name='Aptos Narrow', size=11, color='0563C1', underline='single')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    alt_fill = PatternFill('solid', fgColor='D9E2F3')

    # Write headers
    for col_idx, col_name in enumerate(TEMPLATE_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[1].height = 46

    # Column widths
    widths = [23, 30, 16, 14, 10, 12, 10, 20, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Write data rows with embedded images
    for row_idx, item in enumerate(items, 2):
        ws.row_dimensions[row_idx].height = 100
        fill = alt_fill if row_idx % 2 == 0 else PatternFill(fill_type=None)

        for col_idx, col_name in enumerate(TEMPLATE_COLUMNS, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = item.get(col_name, '')

            if col_name == 'Image' and val:
                img_embedded = False
                img_path = None

                # Check if it's a local extracted image
                if val.startswith('/serve_image/'):
                    local_filename = val.replace('/serve_image/', '')
                    local_path = os.path.join(UPLOAD_FOLDER, secure_filename(local_filename))
                    if os.path.exists(local_path):
                        img_path = local_path

                # Check if it's a remote URL
                elif val.startswith('http'):
                    try:
                        img_data, fmt = download_image(val)
                        if img_data:
                            img_data, fmt = resize_image_bytes(img_data, max_width=120, max_height=100)
                            dl_path = os.path.join(UPLOAD_FOLDER, f'img_{row_idx}.{fmt}')
                            with open(dl_path, 'wb') as f:
                                f.write(img_data)
                            img_path = dl_path
                    except Exception as e:
                        print(f'Image download failed for row {row_idx}: {e}')

                if img_path and os.path.exists(img_path):
                    try:
                        img = XlImage(img_path)
                        img.width = 100
                        img.height = 90
                        cell_ref = f'A{row_idx}'
                        ws.add_image(img, cell_ref)
                        cell.value = ''
                        img_embedded = True
                    except Exception as e:
                        print(f'Image embed failed for row {row_idx}: {e}')

                if not img_embedded:
                    if val.startswith('http'):
                        cell.value = 'View Image'
                        cell.hyperlink = val
                        cell.font = link_font
                    else:
                        cell.value = ''
                        cell.font = cell_font

            elif col_name == 'Retail Link' and val and val.startswith('http'):
                cell.value = 'View Product'
                cell.hyperlink = val
                cell.font = link_font

            elif col_name == 'Cost':
                try:
                    cost_val = float(re.sub(r'[^\d.]', '', str(val))) if val else 0
                    cell.value = cost_val
                    cell.number_format = '$#,##0.00'
                    cell.font = cell_font
                except (ValueError, TypeError):
                    cell.value = val
                    cell.font = cell_font
            else:
                cell.value = val
                cell.font = cell_font

            cell.alignment = cell_align
            cell.border = thin_border
            if fill.fgColor:
                cell.fill = fill

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Save to temp file
    output_path = os.path.join(UPLOAD_FOLDER, 'DLX_Offer_Export.xlsx')
    wb.save(output_path)

    return send_file(output_path, as_attachment=True,
                     download_name='DLX_Distribution_Offer.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=os.environ.get('FLASK_DEBUG', 'false').lower() == 'true')
